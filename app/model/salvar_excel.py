from __future__ import annotations

import unicodedata
from pathlib import Path

import pandas as pd
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


def _calcular_largura(coluna: pd.Series, minimo: int = 12, maximo: int = 40) -> int:
    if coluna.empty:
        return minimo
    comprimento = coluna.astype(str).map(len).max()
    return max(minimo, min(maximo, comprimento + 2))


def _normalizar_nome_coluna(nome: str) -> str:
    texto = unicodedata.normalize("NFKD", nome or "")
    texto = "".join(ch for ch in texto if not unicodedata.combining(ch))
    return texto.lower().strip()


_STATUS_FILLS = {
    "Valores iguais": PatternFill(start_color="92D050", end_color="92D050", fill_type="solid"),
    "Valores divergentes": PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid"),
    "Somente DTE": PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid"),
    "Somente Dominio": PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid"),
    "Somente Domínio": PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid"),
}


def salvar_relatorio(cruzamento: pd.DataFrame, caminho_saida: str | Path) -> None:
    caminho_saida = Path(caminho_saida)
    caminho_saida.parent.mkdir(parents=True, exist_ok=True)

    start_row = 1
    start_col = 1
    header_row = start_row + 1
    first_col_idx = start_col + 1

    currency_columns = {
        "valor dte",
        "valor contabil dominio",
        "diferenca (dominio - dte)",
    }
    integer_columns = {"nota dte", "nota dominio"}

    with pd.ExcelWriter(caminho_saida, engine="openpyxl") as writer:
        cruzamento.to_excel(
            writer,
            sheet_name="Cruzamento",
            index=False,
            startrow=start_row,
            startcol=start_col,
        )
        worksheet = writer.sheets["Cruzamento"]

        total_rows = len(cruzamento)
        total_cols = len(cruzamento.columns)
        last_row = header_row + total_rows
        last_col_idx = first_col_idx + total_cols - 1

        header_fill = PatternFill(start_color="00B0F0", end_color="00B0F0", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        header_alignment = Alignment(horizontal="center", vertical="center")
        central_alignment = Alignment(horizontal="center", vertical="center")
        thin_border = Border(
            left=Side(style="thin", color="000000"),
            right=Side(style="thin", color="000000"),
            top=Side(style="thin", color="000000"),
            bottom=Side(style="thin", color="000000"),
        )

        if total_cols > 0:
            primeira_coluna = get_column_letter(first_col_idx)
            ultima_coluna = get_column_letter(last_col_idx)
            worksheet.auto_filter.ref = f"{primeira_coluna}{header_row}:{ultima_coluna}{last_row}"

        worksheet.row_dimensions[header_row].height = 24
        for col_idx in range(first_col_idx, last_col_idx + 1):
            cell = worksheet.cell(row=header_row, column=col_idx)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
            cell.border = thin_border

        largura_padrao = {
            "Arquivo DTE": 144 / 7,
            "Nota DTE": 96 / 7,
            "Valor DTE": 96 / 7,
            "Chave de Acesso": 308 / 7,
            "Nota Dominio": 126 / 7,
            "Valor Contábil Dominio": 186 / 7,
            "Diferença (Domínio - DTE)": 202 / 7,
            "Status": 119 / 7,
        }

        currency_format = '"R$" #,##0.00;-"R$" #,##0.00;"R$" 0,00'
        integer_format = "#,##0"

        status_col_idx: int | None = None
        for offset, coluna in enumerate(cruzamento.columns):
            col_idx = first_col_idx + offset
            nome_normalizado = _normalizar_nome_coluna(coluna)

            if nome_normalizado == "status":
                status_col_idx = col_idx

            largura = largura_padrao.get(coluna, _calcular_largura(cruzamento[coluna]))
            if nome_normalizado in currency_columns:
                largura = max(largura, 18)

            coluna_letra = get_column_letter(col_idx)
            worksheet.column_dimensions[coluna_letra].width = largura

            for cell in worksheet[coluna_letra]:
                if cell.row == header_row:
                    continue
                cell.border = thin_border
                if nome_normalizado in currency_columns:
                    cell.number_format = currency_format
                    cell.alignment = central_alignment
                elif nome_normalizado in integer_columns:
                    cell.number_format = integer_format
                    cell.alignment = central_alignment
                else:
                    cell.alignment = central_alignment

        # Apply color coding to Status column cells.
        if status_col_idx is not None:
            status_letra = get_column_letter(status_col_idx)
            for row_idx, status_valor in enumerate(cruzamento["Status"], start=header_row + 1):
                fill = _STATUS_FILLS.get(str(status_valor))
                if fill is not None:
                    worksheet.cell(row=row_idx, column=status_col_idx).fill = fill
